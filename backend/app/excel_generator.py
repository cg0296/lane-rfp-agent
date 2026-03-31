import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from app.models import ParsedQuoteData, Carrier


def generate_quote_sheet(
    quote_data: ParsedQuoteData,
    carriers: List[Carrier],
    client_name: str = "Client",
    output_dir: str = "downloads"
) -> tuple[str, str]:
    """
    Generate an Excel quote sheet from parsed data and carriers.

    Args:
        quote_data: Parsed quote data from email
        carriers: List of carriers to quote
        client_name: Name of the client
        output_dir: Directory to save the file

    Returns:
        Tuple of (file_path, filename)
    """

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    info_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    info_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Title
    ws["A1"] = f"Quote Sheet - {client_name}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")

    # Metadata
    row = 3
    ws[f"A{row}"] = "Date Created:"
    ws[f"B{row}"] = datetime.now().strftime("%m/%d/%Y")
    ws[f"A{row}"].font = info_font
    ws[f"B{row}"].fill = info_fill

    row += 1
    ws[f"A{row}"] = "Origin:"
    ws[f"B{row}"] = f"{quote_data.origin_city}, {quote_data.origin_state}"
    ws[f"A{row}"].font = info_font
    ws[f"B{row}"].fill = info_fill

    row += 1
    ws[f"A{row}"] = "Destination:"
    ws[f"B{row}"] = f"{quote_data.destination_city}, {quote_data.destination_state}"
    ws[f"A{row}"].font = info_font
    ws[f"B{row}"].fill = info_fill

    row += 1
    ws[f"A{row}"] = "Equipment:"
    ws[f"B{row}"] = quote_data.equipment_type
    ws[f"A{row}"].font = info_font
    ws[f"B{row}"].fill = info_fill

    row += 1
    ws[f"A{row}"] = "Quantity:"
    ws[f"B{row}"] = quote_data.quantity
    ws[f"A{row}"].font = info_font
    ws[f"B{row}"].fill = info_fill

    if quote_data.loading_date_start:
        row += 1
        ws[f"A{row}"] = "Loading Date:"
        if quote_data.loading_date_end and quote_data.loading_date_start != quote_data.loading_date_end:
            ws[f"B{row}"] = f"{quote_data.loading_date_start} to {quote_data.loading_date_end}"
        else:
            ws[f"B{row}"] = str(quote_data.loading_date_start)
        ws[f"A{row}"].font = info_font
        ws[f"B{row}"].fill = info_fill

    if quote_data.special_requirements:
        row += 1
        ws[f"A{row}"] = "Special Requirements:"
        ws[f"B{row}"] = "; ".join(quote_data.special_requirements)
        ws[f"A{row}"].font = info_font
        ws[f"B{row}"].fill = info_fill

    # Column headers
    row = 11
    headers = ["Lane #", "Origin", "Destination", "Equipment", "Qty"]

    # Add carrier columns
    for carrier in carriers:
        headers.append(carrier.name)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Add data rows (one per quantity unit)
    for lane_num in range(1, quote_data.quantity + 1):
        row += 1
        ws.cell(row=row, column=1).value = lane_num
        ws.cell(row=row, column=2).value = f"{quote_data.origin_city}, {quote_data.origin_state}"
        ws.cell(row=row, column=3).value = f"{quote_data.destination_city}, {quote_data.destination_state}"
        ws.cell(row=row, column=4).value = quote_data.equipment_type

        if quote_data.driver_type:
            ws.cell(row=row, column=4).value += f" ({quote_data.driver_type})"

        ws.cell(row=row, column=5).value = 1

        # Empty cells for carrier rates
        for col_num in range(6, len(headers) + 1):
            cell = ws.cell(row=row, column=col_num)
            cell.alignment = center_align
            cell.border = border

        # Style data row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col_num)
            cell.border = border
            if col_num <= 5:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Adjust column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 8

    for carrier in carriers:
        ws.column_dimensions[get_column_letter(headers.index(carrier.name) + 1)].width = 15

    # Filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"quote_{client_name.replace(' ', '_')}_{timestamp}.xlsx"
    file_path = os.path.join(output_dir, filename)

    # Save
    wb.save(file_path)

    return file_path, filename
